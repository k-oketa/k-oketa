import string

import openpyxl
import pandas as pd

import project


def adjust_dict(result_series: pd.Series, result_format: dict = None) -> dict:
    if result_format is None:
        result_format = {5: None, 4: None, 3: None, 2: None, 1: None}
    results = result_series.to_dict()
    for key in result_format.keys():
        try:
            result_format[key] = results[key]
        except KeyError:
            result_format[key] = 0
    return result_format


if __name__ == '__main__':
    df = pd.DataFrame(columns=['回答者番号', '回答内容'])
    for q_num in range(1, 24):
        q_df = pd.read_csv(project.data_dir + 'D/Q[{}].csv'.format(q_num), header=2, encoding='shift-jis')
        q_df = q_df.rename(columns={'[回答者番号': '回答者番号', ' 回答内容]': '回答内容'})
        q_df = q_df[['回答者番号', '回答内容']]
        df = pd.merge(df, q_df, on='回答者番号', suffixes=('', '_{}'.format(q_num)), how='outer')
    df = df.replace({'強くそう思う': 5, 'ややそう思う': 4, 'どちらとも言えない': 3, 'あまりそう思わない': 2, '全くそう思わない': 1})
    df = df.replace({'５点：': 5, '４点：': 4, '３点：': 3, '２点：': 2, '１点：': 1})
    df = df.replace({'非常に多い': 5, 'やや多い': 4, '適度': 3, 'やや少ない': 2, '非常に少ない': 1})
    df = df.replace({'速すぎる': 5, 'やや速い': 4, '適度': 3, 'やや遅い': 2, '非常に遅い': 1})
    df = df.replace({'10割': 5, '8・9割': 4, '8?9割': 4, '6・7割': 3, '4・5割': 2, '3割以下': 1})
    df = df.replace({'4）上記をmixした講義': 4, '3）Zoom等、リアルタイム双方向型のオンライン講義': 3, '2）ビデオ配信形式のオンライン講義': 2, '1）通常の対面講義': 1})
    df.to_csv(project.data_dir + 'tmp.csv', index_label=False, index=False)

    result_each_question = {}
    for question_number in range(2, 22):
        result = adjust_dict(df['回答内容_{}'.format(question_number)].value_counts())
        result_each_question['回答内容_{}'.format(question_number)] = result

    wb = openpyxl.load_workbook(project.data_dir + 'template.xlsx')
    ws = wb.active
    student_number_cell = ws['H122']
    student_number_cell.value = df['回答者番号'].count()
    alphabets = [alphabet for alphabet in string.ascii_uppercase][3:23]
    for alphabet, (question, result) in zip(alphabets, result_each_question.items()):
        cells = ws['{}127:{}131'.format(alphabet, alphabet)]
        for data, cell in zip(result.values(), cells):
            cell[0].value = data
    wb.save(project.data_dir + 'saved.xlsx')

    q23_result = adjust_dict(df['回答内容_21'].value_counts(), result_format={1: None, 2: None, 3: None, 4: None})
    cells = ws['X162':'X165']
    for data, cell in zip(q23_result.values(), cells):
        cell[0].value = data
    wb.save(project.data_dir + 'saved.xlsx')
